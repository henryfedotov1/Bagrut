from openpyxl import load_workbook
import bcrypt
from constants import file_path

wb = load_workbook(file_path)
ws = wb.active

for row in ws.iter_rows(min_row=2):
    password_cell = row[4]
    plain_password = str(password_cell.value)

    if not plain_password.startswith("$2b$"):
        hashed = bcrypt.hashpw(plain_password.encode(), bcrypt.gensalt()).decode()
        password_cell.value = hashed

wb.save(file_path)
wb.close()

print("הסיסמאות הוצפנו בהצלחה.")
#קטע הקוד הזה ממיר את כל הסיסמאות שיש כרגע שאינן מוצפות לסיסמאות מוצפנות - לא בטוח האם יש צורך להשאיר את זה בפרוקיט כי  אחרי שימוש יחיד הסיסמאות כבר מוצפנות
