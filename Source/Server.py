import socket
import threading
import json
import time
import os
from openpyxl import load_workbook
from constants import file_path, server_port

failed_attempts = {}
blocked_users = {}

def authenticate_user(student_id, password):
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file '{file_path}' not found.")

        wb = load_workbook(file_path)
        if 'Sheet1' not in wb.sheetnames:
            raise ValueError("Sheet isn't found in the Excel file")

        ws = wb['Sheet1']
        headers = [str(cell.value).strip() for cell in ws[1]]

        required_fields = ["ID", "Password", "is_admin", "Paying", "Picture"]
        for field in required_fields:
            if field not in headers:
                raise ValueError(f"Required field '{field}' not found in headers: {headers}")

        id_index = headers.index("ID")
        password_index = headers.index("Password")
        role_index = headers.index("is_admin")
        pay_index = headers.index("Paying")
        picture_index = headers.index("Picture")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[id_index]).strip() == str(student_id).strip():
                stored_password = row[password_index]
                if str(stored_password).strip() == str(password).strip():
                    role = str(row[role_index]).strip().lower()
                    code = 2 if role == "yes" else 1
                    return {
                        'status': 'success',
                        'paying_status': row[pay_index],
                        'picture': row[picture_index],
                        'code': code
                    }

        return {'status': 'fail', 'message': 'Student not found or incorrect password', 'code': 0}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'status': 'fail', 'message': f'Error reading database: {str(e)}', 'code': 0}

def handle_client(client_socket):
    while True:
        try:
            data = client_socket.recv(1024).decode('utf-8')
            if not data:
                break
            request = json.loads(data)
            student_id = request.get('student_id')

            if request['action'] == 'login':
                current_time = time.time()
                if student_id in blocked_users and current_time < blocked_users[student_id]:
                    response = {
                        'status': 'fail',
                        'message': 'Account temporarily locked',
                        'code': 0
                    }
                else:
                    response = authenticate_user(student_id, request['password'])
                    if response['status'] == 'success':
                        failed_attempts[student_id] = 0
                        blocked_users.pop(student_id, None)
                    else:
                        failed_attempts[student_id] = failed_attempts.get(student_id, 0) + 1
                        if failed_attempts[student_id] >= 3:
                            blocked_users[student_id] = time.time() + 60

                client_socket.send(json.dumps(response).encode('utf-8'))
        except:
            break
    client_socket.close()

def start_server():
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.bind(('0.0.0.0', server_port))
    server.listen(5)
    print("[SERVER STARTED] Listening on port 9999...")
    while True:
        client_sock, addr = server.accept()
        thread = threading.Thread(target=handle_client, args=(client_sock,))
        thread.start()

if __name__ == "__main__":
    start_server()
